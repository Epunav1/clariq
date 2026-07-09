"""
Custom data export service for pilot program.
Supports CSV and Excel formats with customizable columns.
"""

import csv
import json
import io
from datetime import datetime
from typing import List, Dict, Optional
from db.pilots_db import list_pilots, get_pilot
from db.actions_db import get_actions, count_actions
from db.revenue_calc import calculate_pilot_revenue
from db.roi_calc import calculate_pilot_roi


# Available columns for export
PILOT_COLUMNS = {
    'id': {'label': 'Pilot ID', 'type': 'int'},
    'name': {'label': 'Pilot Name', 'type': 'str'},
    'email': {'label': 'Email', 'type': 'str'},
    'store_name': {'label': 'Store Name', 'type': 'str'},
    'platform': {'label': 'Platform', 'type': 'str'},
    'status': {'label': 'Status', 'type': 'str'},
    'notes': {'label': 'Notes', 'type': 'str'},
    'created_at': {'label': 'Created At', 'type': 'datetime'},
    'contacted_at': {'label': 'Contacted At', 'type': 'datetime'},
    'completed_at': {'label': 'Completed At', 'type': 'datetime'},
    'days_active': {'label': 'Days Active', 'type': 'int'},
    'reorder_count': {'label': 'Reorders', 'type': 'int'},
    'discount_count': {'label': 'Discounts', 'type': 'int'},
    'promotion_count': {'label': 'Promotions', 'type': 'int'},
    'query_count': {'label': 'Support Queries', 'type': 'int'},
    'total_revenue': {'label': 'Total Revenue', 'type': 'float'},
    'net_profit': {'label': 'Net Profit', 'type': 'float'},
    'roi_percent': {'label': 'ROI %', 'type': 'float'},
    'roi_status': {'label': 'Profitability Tier', 'type': 'str'},
    'cost_per_pilot': {'label': 'Cost Per Pilot', 'type': 'float'},
    'payback_period_days': {'label': 'Payback Period (Days)', 'type': 'int'},
}

ACTION_COLUMNS = {
    'id': {'label': 'Action ID', 'type': 'int'},
    'action_type': {'label': 'Action Type', 'type': 'str'},
    'pilot_id': {'label': 'Pilot ID', 'type': 'int'},
    'pilot_name': {'label': 'Pilot Name', 'type': 'str'},
    'store_id': {'label': 'Store ID', 'type': 'str'},
    'product_id': {'label': 'Product ID', 'type': 'str'},
    'quantity': {'label': 'Quantity', 'type': 'int'},
    'metadata': {'label': 'Metadata', 'type': 'str'},
    'created_at': {'label': 'Created At', 'type': 'datetime'},
}


def get_pilot_export_data(columns: Optional[List[str]] = None) -> List[Dict]:
    """
    Get pilot data formatted for export.
    
    Args:
        columns: List of column names to include. If None, includes all.
    """
    if not columns:
        columns = list(PILOT_COLUMNS.keys())
    
    pilots = list_pilots()
    export_data = []
    
    for pilot in pilots:
        row = {}
        pilot_id = pilot.get('id')
        
        for col in columns:
            if col not in PILOT_COLUMNS:
                continue
            
            if col == 'days_active':
                # Calculate days between contacted and completed/now
                from datetime import datetime
                contacted = pilot.get('contacted_at')
                completed = pilot.get('completed_at') or datetime.utcnow().isoformat()
                if contacted:
                    days = (datetime.fromisoformat(completed) - 
                           datetime.fromisoformat(contacted)).days
                    row[col] = days
            
            elif col in ['reorder_count', 'discount_count', 'promotion_count', 'query_count']:
                # Get action counts
                action_type = col.replace('_count', '')
                row[col] = count_actions(action_type, pilot_id)
            
            elif col in ['total_revenue', 'net_profit', 'roi_percent', 'roi_status', 'cost_per_pilot', 'payback_period_days']:
                # Get ROI data
                try:
                    roi_data = calculate_pilot_roi(pilot_id)
                    if col == 'total_revenue':
                        row[col] = roi_data.get('total_revenue', 0)
                    elif col == 'net_profit':
                        row[col] = roi_data.get('profit_metrics', {}).get('net_profit', 0)
                    elif col == 'roi_percent':
                        row[col] = roi_data.get('roi_metrics', {}).get('roi_percent', 0)
                    elif col == 'roi_status':
                        row[col] = roi_data.get('roi_metrics', {}).get('roi_status', '')
                    elif col == 'cost_per_pilot':
                        row[col] = roi_data.get('investment_metrics', {}).get('cost_per_pilot', 0)
                    elif col == 'payback_period_days':
                        row[col] = roi_data.get('roi_metrics', {}).get('payback_period_days', None)
                except:
                    row[col] = None
            
            else:
                # Direct pilot field
                row[col] = pilot.get(col, '')
        
        export_data.append(row)
    
    return export_data


def get_action_export_data(columns: Optional[List[str]] = None) -> List[Dict]:
    """Get action data formatted for export."""
    if not columns:
        columns = list(ACTION_COLUMNS.keys())
    
    actions = get_actions()
    export_data = []
    
    for action in actions:
        row = {}
        pilot_id = action.get('pilot_id')
        
        for col in columns:
            if col not in ACTION_COLUMNS:
                continue
            
            if col == 'pilot_name' and pilot_id:
                # Lookup pilot name
                try:
                    pilot = get_pilot(pilot_id)
                    row[col] = pilot.get('name', '') if pilot else ''
                except:
                    row[col] = ''
            else:
                row[col] = action.get(col, '')
        
        export_data.append(row)
    
    return export_data


def export_to_csv(data: List[Dict], columns: Optional[List[str]] = None) -> io.StringIO:
    """
    Export data to CSV format.
    
    Args:
        data: List of dictionaries to export
        columns: List of column names. If None, uses all keys from first row
    
    Returns:
        StringIO buffer containing CSV content
    """
    if not data:
        buffer = io.StringIO()
        buffer.write('')
        buffer.seek(0)
        return buffer
    
    if not columns:
        columns = list(data[0].keys())
    
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns, extrasaction='ignore')
    
    writer.writeheader()
    writer.writerows(data)
    
    buffer.seek(0)
    return buffer


def export_to_excel(data: List[Dict], columns: Optional[List[str]] = None, sheet_name: str = 'Data') -> io.BytesIO:
    """
    Export data to Excel format.
    
    Args:
        data: List of dictionaries to export
        columns: List of column names. If None, uses all keys from first row
        sheet_name: Name of the Excel sheet
    
    Returns:
        BytesIO buffer containing Excel file
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not available
        csv_buffer = export_to_csv(data, columns)
        excel_buffer = io.BytesIO(csv_buffer.getvalue().encode('utf-8'))
        excel_buffer.seek(0)
        return excel_buffer
    
    if not data:
        # Return empty workbook
        wb = openpyxl.Workbook()
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    if not columns:
        columns = list(data[0].keys())
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write headers
    header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = row_data.get(col_name, '')
            
            # Format values
            if isinstance(value, (int, float)):
                cell.value = value
                cell.alignment = Alignment(horizontal='right')
            elif isinstance(value, datetime):
                cell.value = value.isoformat()
            else:
                cell.value = str(value) if value is not None else ''
            
            # Alternate row colors
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color='F7F8FB', end_color='F7F8FB', fill_type='solid')
    
    # Auto-fit columns
    for col_idx, col_name in enumerate(columns, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def get_available_columns(data_type: str = 'pilots') -> Dict[str, Dict]:
    """Get list of available columns for a data type."""
    if data_type == 'pilots':
        return PILOT_COLUMNS
    elif data_type == 'actions':
        return ACTION_COLUMNS
    return {}
